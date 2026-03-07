from __future__ import annotations

import zipfile
from io import BytesIO

from anglerfish.deployers.content import render_file_content


def test_render_txt_returns_utf8_bytes_and_text_content_type():
    data, ct = render_file_content("hello world", "notes.txt")

    assert data == b"hello world"
    assert ct == "text/plain; charset=utf-8"


def test_render_docx_returns_valid_docx_bytes():
    data, ct = render_file_content("Test content", "report.docx")

    assert ct == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    # DOCX files are ZIP archives containing OOXML
    assert zipfile.is_zipfile(BytesIO(data))
    with zipfile.ZipFile(BytesIO(data)) as zf:
        assert "[Content_Types].xml" in zf.namelist()


def test_render_xlsx_returns_valid_xlsx_bytes():
    data, ct = render_file_content("Row 1\nRow 2", "data.xlsx")

    assert ct == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    # XLSX files are ZIP archives containing OOXML
    assert zipfile.is_zipfile(BytesIO(data))
    with zipfile.ZipFile(BytesIO(data)) as zf:
        assert "[Content_Types].xml" in zf.namelist()


def test_render_unknown_extension_falls_back_to_txt():
    data, ct = render_file_content("fallback", "file.csv")

    assert data == b"fallback"
    assert ct == "text/plain; charset=utf-8"


def test_render_docx_contains_text_content():
    from docx import Document

    text = "First paragraph\n\nSecond paragraph"
    data, _ = render_file_content(text, "doc.docx")

    doc = Document(BytesIO(data))
    full_text = "\n".join(p.text for p in doc.paragraphs)
    assert "First paragraph" in full_text
    assert "Second paragraph" in full_text


def test_render_xlsx_contains_text_content():
    from openpyxl import load_workbook

    text = "Header Row\nData Row 1\nData Row 2"
    data, _ = render_file_content(text, "sheet.xlsx")

    wb = load_workbook(BytesIO(data))
    ws = wb.active
    assert ws.title == "Data"
    assert ws.cell(row=1, column=1).value == "Header Row"
    assert ws.cell(row=2, column=1).value == "Data Row 1"
    assert ws.cell(row=3, column=1).value == "Data Row 2"


def test_render_empty_text_for_docx():
    data, ct = render_file_content("", "empty.docx")

    assert ct == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    assert zipfile.is_zipfile(BytesIO(data))


def test_render_content_type_mapping():
    _, txt_ct = render_file_content("x", "file.txt")
    _, docx_ct = render_file_content("x", "file.docx")
    _, xlsx_ct = render_file_content("x", "file.xlsx")
    _, unknown_ct = render_file_content("x", "file.pdf")

    assert txt_ct == "text/plain; charset=utf-8"
    assert docx_ct == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    assert xlsx_ct == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert unknown_ct == "text/plain; charset=utf-8"
